from tkinter import *
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
from datetime import datetime
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

root = Tk()
root.geometry("1000x600")
root.title("Bill Management")
root.resizable(False, False)

# Reset all entries to zero
def Reset():
    for item in item_quantities.keys():
        item_quantities[item].set(0)
    Total_bill.set("")

# Calculate the total
def Total():
    totalcost = (
        25 * item_quantities['Samosa'].get() +
        40 * item_quantities['Roll'].get() +
        15 * item_quantities['Pakora'].get() +
        130 * item_quantities['Meatbox'].get() +
        50 * item_quantities['Milkshake'].get() +
        80 * item_quantities['Lassi'].get() +
        100 * item_quantities['Coffee'].get()
    )

    lbl_total = Label(f2, font=("aria", 20, "bold"), text="Total", width=16, fg="lightyellow", bg="black")
    lbl_total.place(x=0, y=50)

    entry_total = Entry(f2, font=("aria", 20, "bold"), textvariable=Total_bill, bd=6, width=15, bg="lightgreen")
    entry_total.place(x=20, y=100)

    # Set the total bill amount in the entry widget
    string_bill = "Tk." + str("%.2f" % totalcost)
    Total_bill.set(string_bill)

# Increment quantity
def increase_quantity(item):
    current_value = item_quantities[item].get()
    item_quantities[item].set(current_value + 1)

# Decrement quantity
def decrease_quantity(item):
    current_value = item_quantities[item].get()
    if current_value > 0:
        item_quantities[item].set(current_value - 1)

# Generate the PDF bill using Table
def print_bill():
    filename = "Bill.pdf"
    doc = SimpleDocTemplate(filename, pagesize=letter)

    current_datetime = datetime.now().strftime("%d-%m-%Y    ; %I:%M:%S %p")

    # Data for the table
    data = [["Item", "Quantity", "Price"]]
    total_cost = 0

    for item, price in menu_items.items():
        quantity = item_quantities[item].get()
        if quantity > 0:
            data.append([item, str(quantity), f"Tk. {price * quantity:.2f}"])
            total_cost += price * quantity

    data.append(["", "Total", f"Tk. {total_cost:.2f}"])

    table = Table(data, colWidths=[200, 100, 150])

    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.white),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.white),
    ]))

    styles = getSampleStyleSheet()
    story = [Paragraph(f"<b>Bill Receipt</b>", styles['Title']),
             Paragraph("<br />", styles['Normal']),
             Paragraph(f"<i>Date: {current_datetime}</i>", styles['Normal']),
             Paragraph("<br />", styles['Normal']),
             table,
             Paragraph("<br />", styles['Normal']),
             Paragraph("<b>******Thanks for purchasing! Have a great day!******</b>", styles['Normal'])]

    try:
        doc.build(story)
        print("Bill saved as Bill.pdf.")
    except Exception as e:
        print(f"Error generating PDF: {e}")

# Export sales data to Excel file
def export_to_excel():
    current_datetime = datetime.now().strftime("%d-%m-%Y %I:%M:%S %p")
    total_cost = 0

    # Gather data for the sale
    sale_data = []
    for item, price in menu_items.items():
        quantity = item_quantities[item].get()
        if quantity > 0:
            sale_data.append(f"{item} x{quantity} (Tk. {price * quantity:.2f})")
            total_cost += price * quantity

    # Create a new Excel workbook or open the existing one
    file_name = "daily_sales.xlsx"
    try:
        wb = openpyxl.load_workbook(file_name)
        sheet = wb.active
    except FileNotFoundError:
        wb = Workbook()
        sheet = wb.active
        # Add headers if file doesn't exist
        headers = ["Sales", "Items Bought", "Total"]
        sheet.append(headers)
        for col in range(1, 4):
            sheet.cell(row=1, column=col).font = Font(bold=True)
            sheet.cell(row=1, column=col).fill = PatternFill(start_color="d2dfae", fill_type="solid")  # Green header
            sheet.cell(row=1, column=col).alignment = Alignment(horizontal="center", vertical="center")

    # Append the sale data to the Excel file
    sheet.append([current_datetime, ", ".join(sale_data), f"Tk. {total_cost:.2f}"])

    # Adjust column widths for better visibility
    sheet.column_dimensions['A'].width = 25
    sheet.column_dimensions['B'].width = 60
    sheet.column_dimensions['C'].width = 15

    # Save the workbook
    wb.save(file_name)
    print("Sale exported to daily_sales.xlsx.")
# GUI Elements
f2 = Frame(root, bg="lightyellow", highlightbackground="black", highlightthickness=1, width=300, height=470)
f2.place(x=690, y=118)

Bill = Label(f2, text="Bill", font=("calibri", 20), bg="lightyellow")
Bill.place(x=120, y=10)

Label(text="BILL MANAGEMENT SYSTEM", bg="black", fg="white", font=("calibri", 33), width="300", height="2").pack()

f = Frame(root, bg="lightblue", highlightbackground="black", highlightthickness=1, width=300, height=470)
f.place(x=10, y=118)

Label(f, text="Menu", font=("Gabriola", 40, "bold"), fg="black", bg="lightblue").place(x=80, y=0)

menu_items = {
    "Samosa": 25,
    "Roll": 40,
    "Pakora": 15,
    "Meatbox": 130,
    "Milkshake": 50,
    "Lassi": 80,
    "Coffee": 100
}

for idx, (item, price) in enumerate(menu_items.items()):
    Label(f, font=("Luca Calligraphy", 15, "bold"), text=f"{item}.......Tk.{price}/pc", fg="black", bg="lightblue").place(x=10, y=80 + idx * 30)

f1 = Frame(root, bd=5, height=375, width=300, relief=RAISED)
f1.pack()

item_quantities = {
    "Samosa": IntVar(value=0),
    "Roll": IntVar(value=0),
    "Pakora": IntVar(value=0),
    "Meatbox": IntVar(value=0),
    "Milkshake": IntVar(value=0),
    "Lassi": IntVar(value=0),
    "Coffee": IntVar(value=0)
}

Total_bill = StringVar()

for idx, item in enumerate(item_quantities.keys()):
    Label(f1, font=("aria", 20, "bold"), text=item, width=8, fg="blue4").grid(row=idx + 1, column=0)
    Button(f1, text="-", font=("aria", 20, "bold"), command=lambda i=item: decrease_quantity(i), width=2).grid(row=idx + 1, column=1)
    entry = Entry(f1, font=("aria", 20, "bold"), textvariable=item_quantities[item], width=3)
    entry.grid(row=idx + 1, column=2)
    Button(f1, text="+", font=("aria", 20, "bold"), command=lambda i=item: increase_quantity(i), width=2).grid(row=idx + 1, column=3)

Button(f1, bd=5, fg="black", bg="lightgreen", font=("ariel", 16, "bold"), width=10, text="Reset", command=Reset).grid(row=8, column=0, pady=20)
Button(f1, bd=5, fg="black", bg="lightgreen", font=("ariel", 16, "bold"), width=10, text="Total", command=Total).grid(row=8, column=2)

Button(f2, bd=5, fg="black", bg="lightblue", font=("ariel", 16, "bold"), width=15, text="Print Bill", command=print_bill).place(x=40, y=300)
Button(f2, bd=5, fg="black", bg="lightblue", font=("ariel", 16, "bold"), width=15, text="Export to CSV", command=export_to_excel).place(x=40, y=380)

root.mainloop()
